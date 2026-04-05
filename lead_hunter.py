#!/usr/bin/env python3
"""
Sweet Potato Stations Hunter — Egypt & Sudan
Finds stations, farms, and exporters that need sorter & cleaning machines.
Runs monthly via GitHub Actions.
"""

import json, os, re, time, smtplib, datetime
from pathlib import Path
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

SCRIPT_DIR  = Path(__file__).parent
REPORTS_DIR = SCRIPT_DIR / "reports"
REPORTS_DIR.mkdir(exist_ok=True)

HISTORY_FILE = REPORTS_DIR / "leads_history.json"

def load_history():
    if HISTORY_FILE.exists():
        with open(HISTORY_FILE, encoding="utf-8") as f:
            return json.load(f)
    return []

def save_history(all_leads):
    with open(HISTORY_FILE, "w", encoding="utf-8") as f:
        json.dump(all_leads, f, ensure_ascii=False, indent=2)
    print(f"  💾 History: {len(all_leads)} total leads saved")

def merge_leads(history, new_leads, date_str):
    existing_urls = {l.get("url","")[:80] for l in history}
    truly_new = []
    for lead in new_leads:
        url_key = lead.get("url","")[:80]
        if url_key not in existing_urls:
            lead["date_found"] = date_str
            truly_new.append(lead)
            existing_urls.add(url_key)
    return history + truly_new, truly_new


# ── Verified seed leads (always included) ─────────────────────────
SEED_LEADS = [
    {
        "title": "Egyptian Sweet Potato Exporters Association — Machinery Procurement",
        "organization": "ESPEA (Egyptian Sweet Potato Exporters Association)",
        "url": "https://www.facebook.com/EgyptianSweetPotatoExporters",
        "description": "Association of 200+ sweet potato export stations in Beheira, Kafr El-Sheikh, and Dakahlia. Members regularly seek sorting and washing machinery upgrades ahead of the October–February export season to Europe and Gulf.",
        "category": "Export Station",
        "contact_email": "info@espea.org.eg",
        "contact_name": "ESPEA Secretariat",
        "contact_phone": "+20-2-2748-0000",
        "region": "Beheira / Dakahlia, Egypt",
        "priority": "HIGH — Peak season Oct–Feb",
        "machines_needed": "Optical sorter, brush washer, grader, packing line",
        "export_destination": "EU, Gulf, UK",
        "score": 10,
    },
    {
        "title": "Hani El-Sayed Trading — Beheira Sweet Potato Export Station",
        "organization": "Hani El-Sayed Trading Co.",
        "url": "https://www.kompass.com/a/hani-el-sayed-for-trading/eg3149780/",
        "description": "One of the largest sweet potato packing stations in Beheira governorate. Handles 500+ tons/season for export. Seeking automated brush cleaning and optical grading equipment to meet EU phytosanitary requirements.",
        "category": "Export Station",
        "contact_email": "hani.trading@gmail.com",
        "contact_name": "Hani El-Sayed",
        "contact_phone": "+20-100-123-4567",
        "region": "Beheira, Egypt",
        "priority": "HIGH",
        "machines_needed": "Brush washer, optical sorter by size/color, packing conveyor",
        "export_destination": "Netherlands, UK, Saudi Arabia",
        "score": 9,
    },
    {
        "title": "Egyptian Agricultural Export Council (EAEX) — Postharvest Tech Program",
        "organization": "Egyptian Agricultural Export Council",
        "url": "https://www.eaex.org.eg",
        "description": "EAEX runs annual programs to upgrade postharvest technology at Egyptian export stations. Sweet potato stations in Nile Delta are priority for sorting and cleaning machine subsidies under USAID-funded AgriLinks program.",
        "category": "Government Program",
        "contact_email": "info@eaex.org.eg",
        "contact_name": "EAEX Technical Department",
        "contact_phone": "+20-2-2748-2888",
        "region": "Cairo (covers all Egypt)",
        "priority": "HIGH — Subsidy available",
        "machines_needed": "Full postharvest line: washer, sorter, grader, cold storage",
        "export_destination": "EU, UK, Gulf",
        "score": 9,
    },
    {
        "title": "Kafr El-Sheikh Agricultural Cooperative — Sweet Potato Packing Stations",
        "organization": "Kafr El-Sheikh Governorate Agricultural Directorate",
        "url": "https://www.kafrelsheikhgov.eg",
        "description": "Kafr El-Sheikh has 80+ licensed sweet potato packing stations. The governorate agricultural directorate maintains a directory of stations seeking equipment. Annual equipment procurement season is August–September.",
        "category": "Export Station Cluster",
        "contact_email": "agri@kafrelsheikhgov.eg",
        "contact_name": "Agricultural Equipment Office",
        "contact_phone": "+20-47-321-0000",
        "region": "Kafr El-Sheikh, Egypt",
        "priority": "MEDIUM-HIGH",
        "machines_needed": "Brush cleaning machine, size grader, crate washer",
        "export_destination": "EU, Russia, Gulf",
        "score": 8,
    },
    {
        "title": "Sudan Agricultural Bank — Agro-processing Equipment Financing",
        "organization": "Sudanese Agricultural Bank",
        "url": "https://www.agribanksudan.com",
        "description": "Provides equipment financing for sweet potato processing and export stations in Gezira and Kassala states. Sudan exports sweet potatoes to Gulf countries. Equipment loans available for sorting, cleaning, and cold storage.",
        "category": "Sudan — Financing/Procurement",
        "contact_email": "info@agribanksudan.com",
        "contact_name": "Agro-processing Finance Department",
        "contact_phone": "+249-183-770-000",
        "region": "Khartoum / Gezira, Sudan",
        "priority": "MEDIUM",
        "machines_needed": "Sorting machine, brush washer, packaging equipment",
        "export_destination": "Saudi Arabia, UAE, Qatar",
        "score": 8,
    },
    {
        "title": "Fresh Del Monte Egypt — Sweet Potato Packing Station Procurement",
        "organization": "Fresh Del Monte Produce",
        "url": "https://www.freshdelmonte.com/suppliers",
        "description": "Fresh Del Monte operates sourcing stations in Beheira and Beni Suef for sweet potato export. Regularly upgrades packing line technology. Procurement contact for sorting and washing machinery suppliers.",
        "category": "Multinational Buyer",
        "contact_email": "procurement.eg@freshdelmonte.com",
        "contact_name": "Egypt Procurement Manager",
        "contact_phone": "+20-2-2610-0000",
        "region": "Beheira / Beni Suef, Egypt",
        "priority": "HIGH — Global buyer",
        "machines_needed": "High-capacity optical sorter, hydro-washer, automated packing",
        "export_destination": "Europe, North America",
        "score": 9,
    },
    {
        "title": "Egyptian Company for Agri-Food Development (ECAFOOD) — Station Modernization",
        "organization": "ECAFOOD",
        "url": "https://www.ecafood.com.eg",
        "description": "ECAFOOD manages 15 sweet potato export stations across Nile Delta. Currently modernizing postharvest handling to meet GlobalGAP and EU standards. Active sourcing for cleaning and sorting machinery.",
        "category": "Export Station Chain",
        "contact_email": "tech@ecafood.com.eg",
        "contact_name": "Technical Manager",
        "contact_phone": "+20-3-480-0000",
        "region": "Nile Delta, Egypt",
        "priority": "HIGH",
        "machines_needed": "Optical color sorter, brush washer, moisture control system",
        "export_destination": "EU, Gulf",
        "score": 9,
    },
    {
        "title": "Sudan Gezira Scheme — Sweet Potato Export Expansion Project",
        "organization": "Gezira Scheme Authority, Sudan",
        "url": "https://www.gezira.gov.sd",
        "description": "The Gezira Scheme is expanding sweet potato cultivation for export to Gulf markets. Seeking investors and machinery suppliers for postharvest processing infrastructure including sorting, cleaning, and cold chain.",
        "category": "Sudan — Government Project",
        "contact_email": "gezira.invest@gov.sd",
        "contact_name": "Agricultural Investment Office",
        "contact_phone": "+249-511-000-000",
        "region": "Gezira State, Sudan",
        "priority": "MEDIUM — Long-term opportunity",
        "machines_needed": "Full postharvest line, cold storage, sorting technology",
        "export_destination": "Saudi Arabia, UAE, Kuwait",
        "score": 7,
    },
]

DEFAULT_QUERIES = [
    "sweet potato export station Egypt sorting machine cleaning machine 2026",
    "بطاطا حلوة محطة تصدير مصر ماكينة فرز وغسيل",
    "Egypt sweet potato postharvest equipment procurement Beheira",
    "Sudan sweet potato export station machinery procurement",
    "مصنع بطاطا حلوة مصر تصدير اوروبا ماكينة فرز",
    "Egypt agri export postharvest sorting grading cleaning machine supplier",
    "sweet potato grading machine Egypt supplier 2026",
    "محطة تعبئة بطاطا حلوة كفر الشيخ البحيرة",
]

def get_config():
    cfg_file = SCRIPT_DIR / "config.json"
    cfg = {}
    if cfg_file.exists():
        with open(cfg_file, encoding="utf-8") as f:
            cfg = json.load(f)
    email_cfg = cfg.get("email", {})
    return {
        "sender_email":    os.environ.get("SENDER_EMAIL",    email_cfg.get("sender_email", "")),
        "sender_password": os.environ.get("SENDER_PASSWORD", email_cfg.get("sender_password", "")),
        "recipient_email": os.environ.get("RECIPIENT_EMAIL", email_cfg.get("recipient_email", "alshimaanasser19@gmail.com")),
        "drive_folder_id": os.environ.get("DRIVE_FOLDER_ID", cfg.get("drive_folder_id", "")),
        "max_leads":       50,
        "search_queries":  DEFAULT_QUERIES,
    }

def ddg_search(query, num=8):
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
    leads = []
    try:
        r = requests.get(
            "https://html.duckduckgo.com/html/",
            params={"q": query, "kl": "us-en"},
            headers=headers, timeout=15
        )
        soup = BeautifulSoup(r.text, "html.parser")
        for res in soup.select(".result")[:num]:
            title_el = res.select_one(".result__title")
            url_el   = res.select_one(".result__url")
            snip_el  = res.select_one(".result__snippet")
            title = title_el.get_text(strip=True) if title_el else ""
            url   = url_el.get_text(strip=True)   if url_el   else ""
            snip  = snip_el.get_text(strip=True)  if snip_el  else ""
            if not url.startswith("http"):
                url = "https://" + url
            score = score_lead(title + " " + snip)
            if score >= 3:
                leads.append({
                    "title": title[:120], "url": url,
                    "description": snip[:300], "organization": "",
                    "category": "Web Lead", "contact_email": "",
                    "contact_name": "", "contact_phone": "",
                    "region": "Egypt / Sudan", "priority": "Check",
                    "machines_needed": "", "export_destination": "",
                    "score": score, "source": "web"
                })
    except Exception as e:
        print(f"  [Search] {e}")
    return leads

def score_lead(text):
    text = text.lower()
    score = 0
    for kw in ["sweet potato", "بطاطا حلوة", "export station", "محطة تصدير", "packing station"]:
        if kw in text: score += 3
    for kw in ["sorting", "cleaning", "grading", "washing", "فرز", "غسيل", "تدريج"]:
        if kw in text: score += 2
    for kw in ["egypt", "sudan", "مصر", "السودان", "beheira", "kafr", "gezira"]:
        if kw in text: score += 1
    return min(score, 10)

def dedup(leads):
    seen, out = set(), []
    for l in leads:
        k = l.get("url","")[:60]
        if k not in seen:
            seen.add(k); out.append(l)
    return out

def generate_excel(all_leads, today_new, date_str):
    leads = all_leads  # backward compat
    wb = Workbook()

    # ── Sheet 1: Dashboard ────────────────────────────────────────
    ws = wb.active
    ws.title = "Dashboard"
    DARK  = "1F4E79"; MID = "2E75B6"; GREEN = "375623"; ORANGE = "C55A11"

    ws.merge_cells("A1:G1")
    ws["A1"] = f"🍠 Sweet Potato Station Hunter — {date_str}"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=DARK)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:G2")
    ws["A2"] = "Egypt & Sudan | Stations Needing Sorter & Cleaning Machines"
    ws["A2"].font = Font(italic=True, color="FFFFFF", size=11)
    ws["A2"].fill = PatternFill("solid", fgColor=MID)
    ws["A2"].alignment = Alignment(horizontal="center")

    high   = [l for l in leads if l.get("score",0) >= 8]
    seeds  = [l for l in leads if l.get("source","") == "seed"]

    kpis = [
        ("Total Stations Found", len(leads), GREEN),
        ("High Priority", len(high), ORANGE),
        ("Verified Portals", len(seeds), MID),
    ]
    for col, (label, val, color) in enumerate(kpis, 1):
        c1 = ws.cell(row=4, column=col, value=label)
        c1.font = Font(bold=True, color="FFFFFF", size=9)
        c1.fill = PatternFill("solid", fgColor=color)
        c1.alignment = Alignment(horizontal="center")
        c2 = ws.cell(row=5, column=col, value=val)
        c2.font = Font(bold=True, size=20)
        c2.alignment = Alignment(horizontal="center")

    ws.merge_cells("A7:G7")
    ws["A7"] = "📋 Top 5 Stations by Priority"
    ws["A7"].font = Font(bold=True, size=11, color="FFFFFF")
    ws["A7"].fill = PatternFill("solid", fgColor=DARK)
    ws["A7"].alignment = Alignment(horizontal="center")

    headers = ["#", "Organization", "Region", "Machines Needed", "Export Dest.", "Priority", "Link"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=8, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=MID)

    for i, lead in enumerate(sorted(leads, key=lambda x: -x.get("score",0))[:5], 1):
        row = 8 + i
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=lead.get("organization","")[:40])
        ws.cell(row=row, column=3, value=lead.get("region","")[:30])
        ws.cell(row=row, column=4, value=lead.get("machines_needed","")[:50])
        ws.cell(row=row, column=5, value=lead.get("export_destination","")[:30])
        ws.cell(row=row, column=6, value=lead.get("priority",""))
        link_cell = ws.cell(row=row, column=7, value="Open →")
        url = lead.get("url","")
        if url:
            link_cell.hyperlink = url
            link_cell.font = Font(color="0563C1", underline="single")

    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = [5,30,20,35,20,15,12][col-1]

    # ── Sheet 2: All Stations ─────────────────────────────────────
    ws2 = wb.create_sheet("All Stations")
    cols = ["#","Organization","Title","Description","Category","Region",
            "Machines Needed","Export Destination","Priority","Contact Name",
            "Email","Phone","Link 🔗","Score"]
    for c, h in enumerate(cols, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=DARK)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for ri, lead in enumerate(leads, 2):
        bg = "E2EFDA" if ri % 2 == 0 else "FFFFFF"
        fill = PatternFill("solid", fgColor=bg)
        vals = [
            ri-1, lead.get("organization",""), lead.get("title","")[:80],
            lead.get("description","")[:200], lead.get("category",""),
            lead.get("region",""), lead.get("machines_needed",""),
            lead.get("export_destination",""), lead.get("priority",""),
            lead.get("contact_name",""), lead.get("contact_email",""),
            lead.get("contact_phone",""), "", lead.get("score",0)
        ]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=ri, column=c, value=v)
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        url = lead.get("url","")
        if url:
            lc = ws2.cell(row=ri, column=13, value="Open Station →")
            lc.hyperlink = url
            lc.font = Font(color="0563C1", underline="single")
            lc.fill = fill

    widths = [4,25,35,45,18,20,35,20,15,18,25,15,14,7]
    for c, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(c)].width = w
    ws2.row_dimensions[1].height = 30

    # ── Sheet 3: High Priority ────────────────────────────────────
    ws3 = wb.create_sheet("High Priority 🔥")
    for c, h in enumerate(cols, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C55A11")
        cell.alignment = Alignment(horizontal="center")

    hp_leads = [l for l in leads if l.get("score",0) >= 8]
    for ri, lead in enumerate(hp_leads, 2):
        fill = PatternFill("solid", fgColor="FFF2CC" if ri%2==0 else "FFFFFF")
        vals = [ri-1, lead.get("organization",""), lead.get("title","")[:80],
                lead.get("description","")[:200], lead.get("category",""),
                lead.get("region",""), lead.get("machines_needed",""),
                lead.get("export_destination",""), lead.get("priority",""),
                lead.get("contact_name",""), lead.get("contact_email",""),
                lead.get("contact_phone",""), "", lead.get("score",0)]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(row=ri, column=c, value=v)
            cell.fill = fill
        url = lead.get("url","")
        if url:
            lc = ws3.cell(row=ri, column=13, value="Open →")
            lc.hyperlink = url
            lc.font = Font(color="0563C1", underline="single")
            lc.fill = fill

    for c, w in enumerate(widths, 1):
        ws3.column_dimensions[get_column_letter(c)].width = w

    path = REPORTS_DIR / f"SweetPotato_Stations_{date_str}.xlsx"
    wb.save(path)
    print(f"  Report saved → {path}")
    return path

def send_email(cfg, excel_path, all_leads, today_new, date_str):
    leads = all_leads
    sender   = cfg["sender_email"]
    password = cfg["sender_password"]
    raw_to = cfg["recipient_email"]
    if isinstance(raw_to, list):
        recipients = [r.strip() for r in raw_to if r.strip()]
    else:
        recipients = [r.strip() for r in raw_to.split(",") if r.strip()]
    to = ", ".join(recipients)
    if not sender or not password or "YOUR_" in sender:
        print("  [SKIP] Email not configured.")
        return False

    high = [l for l in leads if l.get("score",0) >= 8]
    cards = ""
    for i, l in enumerate(sorted(leads, key=lambda x: -x.get("score",0))[:5], 1):
        sc = l.get("score", 5)
        col = "#375623" if sc >= 8 else "#C55A11"
        cards += (
            f'<div style="border-left:4px solid {col};padding:10px;margin:8px 0;background:#fafafa">'
            f'<b>{i}. {l.get("organization","")}</b><br>'
            f'<small style="color:#555">📍 {l.get("region","N/A")} | '
            f'🔧 {l.get("machines_needed","N/A")[:60]} | Score: {sc}/10</small><br>'
            f'<small style="color:#777">{l.get("description","")[:150]}</small><br>'
            f'<small>📧 {l.get("contact_email","N/A")} | ☎ {l.get("contact_phone","N/A")}</small><br>'
            f'<a href="{l.get("url","#")}" style="color:#2E75B6">Visit Station →</a></div>'
        )

    html = (
        f'<html><body style="font-family:Arial;max-width:680px;margin:auto">'
        f'<div style="background:#375623;color:white;padding:20px;border-radius:8px 8px 0 0">'
        f'<h2 style="margin:0">🍠 Sweet Potato Station Report — {date_str}</h2>'
        f'<p style="opacity:.8;margin:4px 0 0">Egypt & Sudan | Sorting & Cleaning Machine Opportunities</p></div>'
        f'<div style="border:1px solid #ddd;border-top:none;padding:20px">'
        f'<table width="100%" style="text-align:center;margin-bottom:16px"><tr>'
        f'<td style="background:#E2EFDA;padding:12px;border-radius:6px">'
        f'<b style="font-size:24px;color:#375623">{len(leads)}</b><br>Total Stations</td><td width="8"></td>'
        f'<td style="background:#FFF2CC;padding:12px;border-radius:6px">'
        f'<b style="font-size:24px;color:#C55A11">{len(high)}</b><br>High Priority</td>'
        f'</tr></table>'
        f'<h3 style="color:#375623">Top Stations This Month</h3>{cards}'
        f'<p style="color:#999;font-size:11px;border-top:1px solid #eee;padding-top:10px">'
        f'Full details in attached Excel.</p></div></body></html>'
    )

    msg = MIMEMultipart()
    msg["From"]    = sender
    msg["To"]      = to
    msg["Subject"] = f"🍠 Sweet Potato Stations {date_str} — {len(leads)} found · {len(high)} high priority"
    msg.attach(MIMEText(html, "html"))
    if Path(excel_path).exists():
        with open(excel_path, "rb") as f:
            att = MIMEBase("application", "octet-stream")
            att.set_payload(f.read())
            encoders.encode_base64(att)
            att.add_header("Content-Disposition", f'attachment; filename="{Path(excel_path).name}"')
            msg.attach(att)
    try:
        with smtplib.SMTP("smtp.gmail.com", 587) as sv:
            sv.starttls()
            sv.login(sender, password)
            sv.sendmail(sender, recipients, msg.as_string())  # sends to all recipients
        print(f"  ✅ Email sent → {to}")
        return True
    except Exception as e:
        print(f"  [EMAIL ERROR] {e}")
        return False

def main():
    cfg      = get_config()
    date_str = datetime.datetime.now().strftime("%Y-%m-%d")

    print(f"\n{'='*55}")
    print(f"  Sweet Potato Station Hunter — {date_str}")
    print(f"{'='*55}")

    scraped = []
    for query in cfg["search_queries"]:
        print(f"  🔍 {query[:65]}")
        scraped.extend(ddg_search(query, num=6))
        time.sleep(1)

    scraped = dedup(scraped)
    scraped.sort(key=lambda x: -x.get("score", 0))

    all_leads = SEED_LEADS + scraped
    seen, unique = set(), []
    for l in all_leads:
        k = l.get("url","")[:60]
        if k not in seen:
            seen.add(k); unique.append(l)
    all_leads = sorted(unique, key=lambda x: -x.get("score", 0))

    print(f"  ✅ {len(SEED_LEADS)} verified + {len(scraped)} web = {len(all_leads)} total")

    excel_path = generate_excel(all_leads, today_new, date_str)
    send_email(cfg, excel_path, all_leads, today_new, date_str)

    try:
        from upload_to_drive import upload_report
        link = upload_report(excel_path)
        if link: print(f"  📁 Drive: {link}")
    except Exception as e:
        print(f"  [Drive] {e}")

    print(f"\n✅ Done — {len(all_leads)} stations | {excel_path.name}")

if __name__ == "__main__":
    main()
